import openpyxl

class excelData():
    def __int__(self,file='data_excel.xlsx'):
        """初始化excel对象"""
        self.file = file
        self.wb=openpyxl.load_workbook(self.file)

    def get_row_value(self,row,sheet_name='Sheet1'):
        sh=self.wb[sheet_name]
        max_col=sh.max_column
        row_value=[]
        for col in range(1,max_col+1):
            value=sh.cell(row,col).value
            row_value.append(value)
        return row_value

    def get_all_row(self,sheet_name='Sheet1'):
        sh= self.wb[sheet_name]
        max_row=sh.max_row
        row_value=[]
        for row in range(2,max_row+1):
            value=self.get_row_value(row)
            row_value.append(value)
        return row_value
if __name__ == '__main__':
    excel=excelData()
    # testdata=excel.get_all_row()
    # print(testdata)
    excel.get_row_value(2)