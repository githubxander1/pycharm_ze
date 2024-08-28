# rows=sh.rows#获取所有行数据
# cells=sh.cell
# write =sh.cell(row=dataInExcel,column=2,value='测试')
# print(write)
# print(rows)
import os

import openpyxl


def get_excel_data3():
    wb = openpyxl.load_workbook('data3_excel.xlsx')
    sh = wb['Sheet1']
    # 2）传入sheet名，获取到sheet页中全部的数据，并转换成list（列表）类型
    # （dataInExcel）通过列表推导式取出excel中第一行的数据，也就是表头
    # title = [i.value for i in res[0]]
    # print(title)
    # （4）循环第二行开始至最后一行的数据，也就是用例数据
    # （5）使用zip进行测试用例表头和用例数据的组装，并转换成dict类型

    # 获取标题内容
    res = list(sh.rows)
    # 遍历除第一行外的其他
    row_data = []
    for item in res[1:]:
        data = [i.value for i in item]
        # dic = dict(zip(data))
        row_data.append(data)
    print(row_data)
    return row_data


# 将最终的测试结果数据，写入excel中。思路：最终结果出来后，复制一份原始的excel文件，
# 写入全部数据后，删除旧的excel表
def write_result(id, result):
    # 打开原始数据文件
    books = openpyxl.load_workbook('data3_excel.xlsx')
    sh1 = books['Sheet1']
    # 复制数据的副本
    sh_names = books.sheetnames
    print(sh_names)
    sh2 = books.copy_worksheet(sh1)
    sh2.title = '含result的表'
    # for shn in books.sheetnames:
    #     if shn.title not in books.sheetnames:
    #         break

    # sheet = wb.get_sheet(0)
    # 测试数据应该与原始行一一对应，所以前面的id（文本格式）作用就体现出来了
    # 把文件中字符串格式的id转换成整数
    id = int(id) + 1
    # 向指定坐标的单元格写入测试数据。
    sh2.cell(row=id, column=8, value=result)  # id：行；7：列
    # 移除原始文件（操作之前，最好备份一份原始测试数据）
    os.remove("data3_excel.xlsx")
    # 保存新文件
    books.save("data3_excel.xlsx")


# 步骤6：
# 最终需要将请求的响应内容拿出去，所以要将结果返回，加return
import requests


def send_request(url, data, data_type, type, **kwargs):
    if type == "post":
        if data_type == "data":
            return requests.post(url=url, data=data, **kwargs)
        elif data_type == "json":
            return requests.post(url=url, json=data, **kwargs)
    elif type == "get":
        return requests.get(url=url, params=data, **kwargs)


if __name__ == '__main__':
    get_excel_data3()
    # write_result(实例25_批量生成PPT版荣誉证书, 'true')
