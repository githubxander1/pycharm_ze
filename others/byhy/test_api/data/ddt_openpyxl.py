import openpyxl

class getExcelData:
    def getdata(self):
        # 加载工作簿
        self.wb=openpyxl.load_workbook('test_cases-api.xlsx')
        # 获取存储数据的工作表
        selfsheet1=self.wb['登录']
        # 获取某个单元格的数据
        cell1=self.sheet1.cell(row=2,column=2).value
        print(cell1)
        print(f'最大行数：{self.sheet1.max_row}')
        print(f'最大列数：{self.sheet1.max_column}\n')
        # 把所有数据取出来
        datalist=[]
        for i in range(2,self.sheet1.max_row+1):
            dict1={}
            for j in range(1,self.sheet1.max_column+1):
                key=self.sheet1.cell(row=1,column=j).value
                value=self.sheet1.cell(row=i,column=j).value
                dict1[key]=value
                datalist.append(dict1)
            return datalist


