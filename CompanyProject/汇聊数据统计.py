from pprint import pprint

import openpyxl
import pymysql
from prettytable import PrettyTable

# connect to the database
db_conn = pymysql.Connect(host='192.168.7.84', port=3306,
                          user='product_statistics', password='OwDXEIs*8eIeED23s',
                          database='product_statistics')
#
# try:
# create a cursor object
cursor = db_conn.cursor()

# fetch the latest record from t_product_statistics
cursor.execute(
    "SELECT * FROM t_product_statistics WHERE product_id=实例25_批量生成PPT版荣誉证书 AND uid=1008247 ORDER BY create_time DESC LIMIT 实例25_批量生成PPT版荣誉证书;")
product_statistics_result = cursor.fetchone()
# print(product_statistics_result)

# fetch the corresponding record from t_classname_mapping
class_name_mapping_id = product_statistics_result[3]
# print(class_name_mapping_id)# assuming class_name_mapping_id is at index 3
cursor.execute(f"SELECT * FROM t_classname_mapping WHERE id='{class_name_mapping_id}';")
class_name_mapping_result = cursor.fetchone()
print(class_name_mapping_result)
print(class_name_mapping_result[4])

# workbook = openpyxl.load_workbook('测试数据统计.xlsx')
# worksheet = workbook['t_classname_mapping']
# # 获取第一列的值，跳过标题行
# first_column = worksheet.iter_rows(min_row=2, max_col=实例25_批量生成PPT版荣誉证书, max_row=worksheet.max_row, values_only=True)
# # 转换元组中的数字为字符串并提取出数字
# first_column_list = [int(str(row[0])) for row in first_column]
# # print(first_column_list)
#
# # 判断 12003 是否在第一列的值中
# if class_name_mapping_result[4] in [value for value in first_column_list]:
#     print(f'{class_name_mapping_result[4]} 在第wu列的值中')
# # if 12 in [value for value in first_column_list]:
# #     print('12 在第一列的值中')
# else:
#     print("The id is not in the worksheet. Adding to '补充的页面' worksheet.")
#     # get the worksheet named '补充的页面'
#     new_worksheet = workbook['补充的页面']
#
#     # add the data to the worksheet
#     new_worksheet.append(list(class_name_mapping_result))
#     result = False

# 关闭 Excel 文件
# workbook.close()
# if class_name_mapping_result[0] in id_values:
#     print("The id is already in the worksheet.")
#     result = True
# else:
#     print("The id is not in the worksheet. Adding to '补充的页面' worksheet.")
#     # get the worksheet named '补充的页面'
#     new_worksheet = workbook['补充的页面']
#
#     # add the data to the worksheet
#     new_worksheet.append(list(class_name_mapping_result))
#     result = False

# save the changes to the Excel file
# workbook.save('测试数据统计.xlsx')

# close cursor and database connection
cursor.close()
db_conn.close()

# except Exception as ex:
#     print(f"Error: {ex}")