import os
from openpyxl import load_workbook

# 使用os.path.join处理文件路径，确保跨平台兼容性
file_path = os.path.join("data", "领料单（每日）.xlsx")
output_file_path = os.path.join("data", "领料单（每日）-更.xlsx")

try:
    # 使用try-except结构来处理可能发生的异常，如文件不存在或损坏
    wb = load_workbook(file_path)
except Exception as e:
    print(f"打开Excel文件时出错：{e}")
    exit(1)

sheet_names = wb.sheetnames

for sheet_name in sheet_names:
    try:
        # 遍历每个工作表，更改A4单元格的数据
        ws = wb[sheet_name]
        ws['A4'].value = "零件测试领料单1"
    except Exception as e:
        print(f"处理工作表 {sheet_name} 时出错：{e}")

# # 在保存前检查输出文件是否已经存在，以避免覆盖原有文件
# if os.path.exists(output_file_path):
#     print(f"输出文件 {output_file_path} 已经存在！")
#     # 你可以在这里添加一个确认步骤，询问用户是否覆盖文件
# else:
#     try:
#         # 保存更改后的文件
#         wb.save(output_file_path)
#         print(f"文件已成功保存至：{output_file_path}")
#     except Exception as e:
#         print(f"保存Excel文件时出错：{e}")
