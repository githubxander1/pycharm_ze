

# os
import os
# import pprint
import shutil
from pprint import pprint

print(os.getcwd())
# pprint(os.listdir())
pprint(i for i in os.walk(os.getcwd()))

dirpath ='testdir'

# def remove_dir(dirpath):
#     if os.path.exists(dirpath) and os.path.isdir(dirpath):
#         try:
#             shutil.rmtree(dirpath)
#             print('rmtree成功')
#         except Exception as e:
#             print(e)
#     else:
#         print('目录不存在或不是目录')
#
# def creat_dir(dirpath):
#     if not os.path.exists(dirpath):
#         try:
#             os.makedirs(dirpath)
#             print('创建目录成功')
#         except Exception as e:
#             print(e)
#
# # os.makedirs('textdir/dir1/dir2')
# print(os.path.basename(dirpath))

from openpyxl import Workbook, load_workbook


def creat_worksheet(wb,sheet_name,index):
    '''
    创建并返回工作表
    '''
    ws = wb.creat_sheet(sheet_name,index)
    ws.sheet_properties.tabColor = '107'
    return ws
def write_data_to_sheet(ws,data):
    '''
    向指定工作表写入数据
    '''
    for row in data:
        ws.append(row)
wb1 = Workbook('flashing_icon.xlsx')
wb= load_workbook('test.xlsx')
print(wb.sheetnames)
ws = wb['Sheet']
for row in ws.iter_rows(min_row=1,max_row=3,min_col=1,max_col=3):
    for data in row:
        print(data.value)